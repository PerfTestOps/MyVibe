import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def sync_forecasts_to_sheet1_from_df(excel_file, sheet2_df):
    sheet1 = pd.read_excel(excel_file, sheet_name="Sheet1", engine="openpyxl")

    # Normalize IDs and project names
    sheet1["AssociateID"] = sheet1["AssociateID"].astype(str).str.strip()
    sheet1["Project Name"] = sheet1["Project Name"].astype(str).str.strip()
    sheet2_df["AssociateID"] = sheet2_df["AssociateID"].astype(str).str.strip()
    sheet2_df["Project Name"] = sheet2_df["Project Name"].astype(str).str.strip()

    # Detect valid months based on available columns
    month_prefixes = ["Jan", "Feb", "Mar", "April", "May", "Jun",
                      "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    months = [m for m in month_prefixes if any(col.startswith(m + "-") for col in sheet2_df.columns)]

    # Map column names from Sheet2 and Sheet1
    sheet2_available = {m: next((c for c in sheet2_df.columns if c.startswith(f"{m}-Available")), None) for m in months}
    sheet2_leave = {m: next((c for c in sheet2_df.columns if c.startswith(f"{m}-Leave")), None) for m in months}
    sheet2_extra = {m: next((c for c in sheet2_df.columns if c.startswith(f"{m}-Extra")), None) for m in months}
    sheet1_forecast = {m: next((c for c in sheet1.columns if f"{m} Forecast" in c), None) for m in months}

    # Build forecast data
    forecast_map = {}
    for _, row in sheet2_df.iterrows():
        aid = row["AssociateID"]
        pname = row["Project Name"]
        billing_rate = pd.to_numeric(row.get("BillingRateByMonth", 0), errors="coerce") or 0

        forecast_map.setdefault(aid, {})[pname] = {}
        for m in months:
            a = pd.to_numeric(row.get(sheet2_available[m], 0), errors="coerce") or 0
            l = pd.to_numeric(row.get(sheet2_leave[m], 0), errors="coerce") or 0
            e = pd.to_numeric(row.get(sheet2_extra[m], 0), errors="coerce") or 0
            forecast = (a + e - l) * billing_rate

            target_col = sheet1_forecast[m]
            if target_col:
                forecast_map[aid][pname][target_col] = forecast

    # Apply forecast to Sheet1
    for idx, row in sheet1.iterrows():
        aid = row["AssociateID"]
        pname = row["Project Name"]
        if aid in forecast_map and pname in forecast_map[aid]:
            for col, val in forecast_map[aid][pname].items():
                if col in sheet1.columns:
                    sheet1.at[idx, col] = val

    # Overwrite Sheet1
    wb = load_workbook(excel_file)
    if "Sheet1" in wb.sheetnames:
        wb.remove(wb["Sheet1"])
    ws = wb.create_sheet("Sheet1")
    ws.append(sheet1.columns.tolist())
    for r in sheet1.itertuples(index=False):
        ws.append(list(r))
    wb.save(excel_file)

def show_page():
    excel_file = "BaseDatasheet.xlsx"
    sheet_name = "Sheet2"

    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")
    df["AssociateID"] = df["AssociateID"].astype(str).str.strip()
    df["Project Name"] = df["Project Name"].astype(str).str.strip()

    st.title("🧠 Update Forecast Data")

    # 🔍 Filter by Project Name
    project_filter = st.text_input("Filter by Project Name (Exact Match)", placeholder="Type exact project name")
    filtered_df = df[df["Project Name"] == project_filter] if project_filter else df.copy()

    # 🔐 Role check
    role = st.session_state.get("role", "Unknown")

    if role == "Viewer":
        # 📋 View-only table
        st.info("🔒 Read-only access: Forecast data can be viewed but not edited.")
        st.dataframe(filtered_df, use_container_width=True)
    else:
        # ✏️ Editable table for Admin/Contributor
        edited_df = st.data_editor(filtered_df, num_rows="fixed", use_container_width=True)

        # 💾 Save and Sync Forecasts
        if st.button("💾 Save Changes and Sync Forecasts"):
            try:
                wb = load_workbook(excel_file)
                ws = wb[sheet_name]

                if project_filter:
                    updated_df = df.copy()
                    match_indices = updated_df[updated_df["Project Name"] == project_filter].index.tolist()
                    for i, idx in enumerate(match_indices):
                        updated_df.loc[idx] = edited_df.iloc[i]
                else:
                    updated_df = edited_df.copy()

                # Overwrite Sheet2
                wb.remove(ws)
                ws = wb.create_sheet(sheet_name)
                ws.append(updated_df.columns.tolist())
                for r in dataframe_to_rows(updated_df, index=False, header=False):
                    ws.append(r)
                wb.save(excel_file)

                # Sync forecasts to Sheet1
                sync_forecasts_to_sheet1_from_df(excel_file, updated_df)

                st.success("✅ Forecast data saved and synced to Sheet1 successfully.")
            except Exception as e:
                st.error(f"❌ Save or sync failed: {e}")
