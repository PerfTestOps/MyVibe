import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
import re

EXCEL_PATH = "BaseDatasheet.xlsx"

def load_data(file_path: str, sheet_name: str = "Sheet2") -> pd.DataFrame:
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df.columns = [c.strip() for c in df.columns]
    df['ServiceLine'] = df['ServiceLine'].astype(str).str.strip().str.upper()
    df['PracticeLine'] = df['PracticeLine'].astype(str).str.strip()
    df['Region'] = df['Region'].astype(str).str.strip()
    return df

def apply_filters(df: pd.DataFrame) -> pd.DataFrame:
    st.sidebar.title("🔎 Filter By")
    filters = {
        'ServiceLine': st.sidebar.multiselect("Service Line", df['ServiceLine'].dropna().unique()),
        'PracticeLine': st.sidebar.multiselect("Practice Line", df['PracticeLine'].dropna().unique()),
        'Region': st.sidebar.multiselect("Region", df['Region'].dropna().unique())
    }
    for key, selected in filters.items():
        if selected:
            df = df[df[key].isin(selected)]
    return df

def extract_months(df: pd.DataFrame) -> list:
    pattern = re.compile(r'^([A-Za-z]+)-Available Days$')
    months = []
    for col in df.columns:
        match = pattern.match(col)
        if match:
            months.append(match.group(1))
    return months

def calculate_forecast(df: pd.DataFrame, months: list) -> pd.DataFrame:
    forecast_df = []
    for month in months:
        av_col = f"{month}-Available Days"
        lv_col = f"{month}-Leave Days"
        ex_col = f"{month}-Extra Working Days"

        if all(c in df.columns for c in [av_col, lv_col, ex_col]):
            df['Forecast'] = (df[av_col] + df[ex_col] - df[lv_col]) * df['BillingRateByMonth']
            df['Month'] = month
            forecast_df.append(df[['AssociateName', 'ServiceLine', 'PracticeLine', 'Region', 'Month', 'Forecast']])
    return pd.concat(forecast_df, ignore_index=True)

def plot_chart(df: pd.DataFrame):
    grouped = df.groupby(['Month', 'ServiceLine'], as_index=False)['Forecast'].sum()
    fig = px.bar(
        grouped,
        x="Month",
        y="Forecast",
        color="ServiceLine",
        barmode="group",
        text="Forecast",
        title="📊 Forecast Revenue by Month",
        height=500
    )
    fig.update_traces(textposition='outside')
    st.plotly_chart(fig, use_container_width=True)

def update_excel_runtime(forecast_df: pd.DataFrame, sheet_name: str = "Sheet1"):
    wb = load_workbook(EXCEL_PATH)
    if sheet_name not in wb.sheetnames:
        st.error(f"Sheet '{sheet_name}' not found.")
        return

    sheet = wb[sheet_name]
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1) if cell.value}
    for row in range(2, sheet.max_row + 1):
        ass_name = sheet.cell(row=row, column=headers.get("AssociateName")).value
        serv_line = sheet.cell(row=row, column=headers.get("ServiceLine")).value
        prac_line = sheet.cell(row=row, column=headers.get("PracticeLine")).value
        region = sheet.cell(row=row, column=headers.get("Region")).value

        for _, row_df in forecast_df.iterrows():
            if (
                row_df['AssociateName'] == ass_name and
                row_df['ServiceLine'] == serv_line and
                row_df['PracticeLine'] == prac_line and
                row_df['Region'] == region
            ):
                month_col = f"{row_df['Month']} Forecast"
                if month_col in headers:
                    col_idx = headers[month_col]
                    sheet.cell(row=row, column=col_idx).value = round(row_df['Forecast'])

    wb.save(EXCEL_PATH)
    st.success(f"✅ Excel updated with forecast in '{sheet_name}'.")

def main():
    st.title("📈 Practice Forecast Calculator")

    df = load_data(EXCEL_PATH)
    df_filtered = apply_filters(df)
    months = extract_months(df_filtered)
    forecast_df = calculate_forecast(df_filtered, months)

    if forecast_df.empty:
        st.warning("⚠️ No data available after filtering.")
    else:
        plot_chart(forecast_df)
        update_excel_runtime(forecast_df)
        # if st.button("📥 Update Forecast in Sheet1"): # if need to update on button then enable this and move update_excel_runtime inside this

if __name__ == "__main__":
    main()