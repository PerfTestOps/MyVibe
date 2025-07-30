import streamlit as st
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

def show_page():
    st.title("👥 Add Associate to Revenue Tracker")

    # 🔐 Role restriction
    if st.session_state.get("role") == "Viewer":
        st.warning("🚫 You don't have permission to add users. This section is restricted.")
        st.stop()

    excel_file = "BaseDatasheet.xlsx"

    dropdown_options = {
        "vertical": ["BFS", "Insurance", "CMT", "HCLS", "Internal","PnR"],
        "sub_vertical": ["BFS", "Communication", "GPS", "Health Care", "Life Science","MLEU","RCG","T&H","Technology"],
        "region": ["Americas", "APAC", "APJ", "EMEA","GGM"],
        "sub_region": ["ANZ", "APAC", "ASEAN", "DACH", "GGM", "Germany","India","Japan","NA","Northern Europe","Southern Europe & Middle East","Switzerland","UK","UK&I"],
        "service_line": ["AT", "PE", "PT", "SDET", "SRE", "ST", "UT"],
        "bl_ml": ["BL", "ML"],
        "specifics": ["Frontend", "Backend", "Fullstack", "Test Automation", "Data Analytics"],
        "project_type": ["Managed Service", "Staff Augmentation", "Development", "QA", "T&M",   
                       "Fixed Price", "Support","Consulting", "Training","FBT", "Transaction Based"],
        "onshore_offshore": ["Onshore", "Offshore"],
        "digital": ["Digital", "Non-Digital"],
        "digital_category": ["Experience", "Intelligence", "Engineering", "Modernization"],
        "methodology": ["Agile", "Waterfall", "Hybrid", "Scrum", "Kanban"]
    }

    if not os.path.exists(excel_file):
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Sheet1"
        header = [
            "Vertical", "Sub Vertical", "Parent Customer", "Project Description",
            "Region", "Sub Region", "Associate ID", "Associate Name",
            "Onshore/Offshore", "Service Line", "Methodology", "BL/ML",
            "Digital/Non-Digital", "Digital Category", "Specifics", "Project Type",
            "Demand Start Date", "Resource Start Date", "Hourly Rate (USD)",
            "Day Rate (USD)", "Entry Timestamp"
        ]
        sheet1.append(header)
        sheet2 = wb.create_sheet("Sheet2")
        sheet2.append(header)
        wb.save(excel_file)

    wb = load_workbook(excel_file)
    sheet1 = wb["Sheet1"]
    sheet2 = wb["Sheet2"]

    with st.form("add_user_form"):
        col1, col2 = st.columns(2)

        with col1:
            vertical = st.selectbox("Vertical", dropdown_options["vertical"])
            sub_vertical = st.selectbox("Sub Vertical", dropdown_options["sub_vertical"])
            parent_customer = st.text_input("Parent Customer")
            project_description = st.text_input("Project Description")
            region = st.selectbox("Region", dropdown_options["region"])
            sub_region = st.selectbox("Sub Region", dropdown_options["sub_region"])
            associate_id = st.text_input("Associate ID")
            associate_name = st.text_input("Associate Name")
            onshore_offshore = st.selectbox("Onshore/Offshore", dropdown_options["onshore_offshore"])
            service_line = st.selectbox("Service Line", dropdown_options["service_line"])

        with col2:
            methodology = st.selectbox("Methodology", dropdown_options["methodology"])
            bl_ml = st.selectbox("BL/ML", dropdown_options["bl_ml"])
            digital = st.selectbox("Digital/Non-Digital", dropdown_options["digital"])
            digital_category = st.selectbox("Digital Category", dropdown_options["digital_category"])
            specifics = st.selectbox("Specifics", dropdown_options["specifics"])
            project_type = st.selectbox("Project Type", dropdown_options["project_type"])
            demand_start = st.date_input("Demand Start Date")
            resource_start = st.date_input("Resource Start Date")
            hourly_rate = st.number_input("Hourly Rate (USD)", min_value=0.0)
            day_rate = st.number_input("Day Rate (USD)", min_value=0.0)

        submitted = st.form_submit_button("Submit")

    if submitted:
        existing_ids = [str(row[6].value) for row in sheet1.iter_rows(min_row=2) if row[6].value]

        if associate_id in existing_ids:
            st.warning(f"⚠️ Associate ID '{associate_id}' already exists. Entry not added.")
        else:
            row_data = [
                vertical, sub_vertical, parent_customer, project_description,
                region, sub_region, associate_id, associate_name,
                onshore_offshore, service_line, methodology, bl_ml,
                digital, digital_category, specifics, project_type,
                str(demand_start), str(resource_start), hourly_rate,
                day_rate, datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]
            sheet1.append(row_data)
            sheet2.append(row_data)
            wb.save(excel_file)
            st.success(f"✅ Associate '{associate_name}' added successfully to both sheets!")
