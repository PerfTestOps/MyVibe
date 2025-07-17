import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import os

# Load Excel data into a DataFrame
excel_file = "BaseDatasheet.xlsx"
sheet_name = "Sheet1"

# Read the Excel file
df = pd.read_excel(excel_file, sheet_name=sheet_name)

st.title("Updating Actuals / Forecast")

# Show editable table
edited_df = st.data_editor(df, num_rows="fixed", use_container_width=True)

# Save changes back to Excel
if st.button("Save Changes"):
    # Load workbook and overwrite the sheet with edited data
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Clear existing data
    ws.delete_rows(2, ws.max_row)

    # Write updated rows
    for r_idx, row in edited_df.iterrows():
        ws.append(row.tolist())

    wb.save(excel_file)
    st.success("Changes saved to Excel!")
