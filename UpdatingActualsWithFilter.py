import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def show_page():
    excel_file = "BaseDatasheet.xlsx"
    sheet_name = "Sheet1"
    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine="openpyxl")

    st.title("🧠 Update Actuals Data")

    # 🔍 Filter by Project Name
    project_filter = st.text_input("Filter by Project Name (Exact Match)", placeholder="Type exact project name")

    df["AssociateID"] = df["AssociateID"].astype(str)
    filtered_df = df[df["Project Name"] == project_filter] if project_filter else df.copy()

    role = st.session_state.get("role", "Unknown")

    if role == "Viewer":
        st.info("🔒 Read-only access: You can view data but cannot edit.")
        st.dataframe(filtered_df, use_container_width=True)
    else:
        edited_df = st.data_editor(filtered_df, num_rows="fixed", use_container_width=True)

        if st.button("💾 Save Changes"):
            try:
                wb = load_workbook(excel_file)
                ws = wb[sheet_name]

                header = [cell.value for cell in ws[1]]
                project_col_index = header.index("Project Name")

                if project_filter:
                    updated_df = df.copy()
                    matched_indices = updated_df[updated_df["Project Name"] == project_filter].index.tolist()
                    for i, row_idx in enumerate(matched_indices):
                        updated_df.loc[row_idx] = edited_df.iloc[i]
                else:
                    updated_df = edited_df.copy()

                wb.remove(ws)
                ws = wb.create_sheet(sheet_name)
                ws.append(updated_df.columns.tolist())
                for r in dataframe_to_rows(updated_df, index=False, header=False):
                    ws.append(r)

                wb.save(excel_file)
                st.success(f"✅ Changes saved to Sheet1{' for project: ' + project_filter if project_filter else ''}")
            except Exception as e:
                st.error(f"❌ Failed to save changes: {e}")
