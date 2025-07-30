import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# --- Security Check ---
def is_admin():
    return st.session_state.get("role") == "Admin"

# --- Excel Configuration ---
excel_path = "BaseDatasheet.xlsx"
target_sheet = "AdminWorks"

# --- Add Project Row ---
def add_project_to_adminworks(parent_customer, project_description):
    book = load_workbook(excel_path)

    if target_sheet not in book.sheetnames:
        raise Exception(f"Sheet '{target_sheet}' not found in workbook.")

    ws = book[target_sheet]
    headers = [cell.value for cell in ws[1]]
    row = [""] * len(headers)

    if "Parent customer" in headers:
        idx = headers.index("Parent customer")
        row[idx] = parent_customer
    if "Project description" in headers:
        idx = headers.index("Project description")
        row[idx] = project_description

    ws.append(row)
    book.save(excel_path)

# --- Read Sheet as DataFrame ---
def load_adminworks_df():
    return pd.read_excel(excel_path, sheet_name=target_sheet, engine="openpyxl")

# --- Delete Row by Index ---
def delete_entry(row_index):
    book = load_workbook(excel_path)
    ws = book[target_sheet]
    ws.delete_rows(row_index + 2)  # +2 accounts for header row + 0-based index
    book.save(excel_path)

# --- Main UI Page ---
def show_page():
    st.title("🛠️ Admin: Add or Manage Projects")
    st.markdown("---")

    if not is_admin():
        st.warning("🚫 You do not have permission to access this page.")
        st.stop()

    # --- Add New Project Form ---
    with st.form("add_project_form"):
        parent_customer = st.text_input("Parent Customer")
        project_description = st.text_area("Project Description", height=100)
        submitted = st.form_submit_button("➕ Add Project")

    if submitted:
        try:
            add_project_to_adminworks(parent_customer, project_description)
            st.success("✅ Project added to AdminWorks successfully.")
        except Exception as e:
            st.error(f"⚠️ Failed to add project: {e}")

    st.markdown("### 📄 Current Entries in AdminWorks")
    try:
        df = load_adminworks_df()
        if df.empty:
            st.info("No entries found.")
        else:
            st.dataframe(df, use_container_width=True)

            st.markdown("### 🔧 Delete a Project Entry")
            with st.form("delete_form"):
                selected_index = st.number_input(
                    "Enter row number to delete (starting from 0)",
                    min_value=0,
                    max_value=len(df) - 1,
                    step=1
                )
                confirm_delete = st.form_submit_button("🗑️ Delete Entry")

            if confirm_delete:
                try:
                    delete_entry(selected_index)
                    st.success(f"✅ Entry at row {selected_index} deleted successfully.")
                except Exception as e:
                    st.error(f"⚠️ Failed to delete entry: {e}")
    except Exception as e:
        st.error(f"⚠️ Unable to read sheet: {e}")
